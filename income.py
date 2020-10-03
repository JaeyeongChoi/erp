import openpyxl

def income(filename, filename2):
    #    filename = "aa.xlsx"
    #    filename2 = "bb.xlsx"
    wb = openpyxl.load_workbook(filename)
    wb2 = openpyxl.load_workbook(filename2)
    sheet = wb["Sheet1"]
    sheet2 = wb2["Sheet1"]
    i=1
    while(sheet.cell(i,1).value):
        sheet2.cell(i,1).value = "test"
        i=i+1
#use wb.get_highest_row()
#    sheet2['A1'].value = sheet['A1'].value
    #sheet.title = "testtest"

    wb.save(filename)
    wb2.save(filename2)


def excelcopy(sheet1, sheet2, type):

    if(type == 1):
        sheet1.cell(1,1).value = "test"






filename = "aa.xlsx"
filename2 = "bb.xlsx"
income(filename, filename2)
